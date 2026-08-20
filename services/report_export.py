"""Generate durable class and individual teaching reports from history data."""

from __future__ import annotations

import html
import re
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from models.history import HistoricalStudent, HistoricalTask
from repositories.database import default_data_directory
from services.result_query import project_historical_results


class ReportExportError(RuntimeError):
    """Base class for report generation failures."""


class ExportPermissionError(ReportExportError):
    """Raised when the export directory cannot be written."""


def export_class_report(
    task: HistoricalTask, *, output_dir: str | Path | None = None
) -> Path:
    directory = _prepare_directory(output_dir)
    path = directory / f"{_report_stem(task)}.xlsx"
    rows = project_historical_results(task)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "总体统计"
    students = workbook.create_sheet("学生结果")
    diagnoses = workbook.create_sheet("错误类型统计")

    status_counts = Counter(row.status for row in rows)
    summary.append(["班级诊断报告"])
    summary.append(["任务名称", task.task.title])
    summary.append(["题目 ID", task.task.problem_id])
    summary.append(["创建时间", task.task.created_at.astimezone().replace(tzinfo=None)])
    summary.append(["总人数", len(rows)])
    summary.append(["AC 人数", status_counts.get("AC", 0)])
    summary.append(["WA 人数", status_counts.get("WA", 0)])
    summary.append(["TLE 人数", status_counts.get("TLE", 0)])
    summary.append(["AC 率", status_counts.get("AC", 0) / len(rows) if rows else 0])
    summary["B4"].number_format = "yyyy-mm-dd hh:mm"
    summary["B9"].number_format = "0%"

    students.append(
        ["学生姓名", "状态", "通过数", "总测试点", "通过率", "主要诊断", "AI解释状态"]
    )
    for row in rows:
        students.append(
            [
                row.student_name,
                row.status,
                row.passed_count,
                row.total_count,
                row.pass_rate,
                row.diagnosis_category or "-",
                row.explanation_status,
            ]
        )
    for cell in students["E"][1:]:
        cell.number_format = "0%"
    if students.max_row > 1:
        students.auto_filter.ref = f"A1:G{students.max_row}"
        students.freeze_panes = "A2"
        students.conditional_formatting.add(
            f"E2:E{students.max_row}",
            CellIsRule(operator="lessThan", formula=["1"], fill=PatternFill("solid", fgColor="FDE2E2")),
        )

    diagnosis_counts = Counter(
        row.diagnosis_category for row in rows if row.diagnosis_category
    )
    diagnoses.append(["错误类型", "人数"])
    for category, count in sorted(diagnosis_counts.items()):
        diagnoses.append([category, count])
    if not diagnosis_counts:
        diagnoses.append(["无规则诊断", 0])

    for sheet in workbook.worksheets:
        _style_sheet(sheet)
    try:
        workbook.save(path)
    except PermissionError as exc:
        raise ExportPermissionError("无法写入导出目录。") from exc
    except OSError as exc:
        raise ReportExportError("报告生成失败。") from exc
    return path


def export_student_feedback(
    task: HistoricalTask,
    submission_id: str,
    *,
    output_dir: str | Path | None = None,
) -> Path:
    student = next(
        (item for item in task.students if item.submission.id == submission_id), None
    )
    if student is None:
        raise ReportExportError("学生记录不存在。")
    directory = _prepare_directory(output_dir)
    path = directory / (
        f"{_report_stem(task)}_{_safe_filename(student.submission.student_name)}.html"
    )
    content = _student_feedback_html(task, student)
    try:
        path.write_text(content, encoding="utf-8", newline="\n")
    except PermissionError as exc:
        raise ExportPermissionError("无法写入导出目录。") from exc
    except OSError as exc:
        raise ReportExportError("报告生成失败。") from exc
    return path


def _student_feedback_html(task: HistoricalTask, student: HistoricalStudent) -> str:
    submission = student.submission
    diagnosis = student.diagnosis
    explanation = student.explanation
    analysis = diagnosis.detail if diagnosis and diagnosis.detail else "暂无明确规则诊断。"
    suggestion = (
        explanation.student_explanation
        if explanation and explanation.student_explanation
        else "请结合失败测试点检查代码逻辑和边界条件。"
    )
    evidence = "".join(
        f"<li>{html.escape(item)}</li>" for item in (diagnosis.evidence if diagnosis else ())
    ) or "<li>暂无证据</li>"
    return f"""<!doctype html>
<html lang="zh-CN"><head><meta charset="utf-8"><title>学生反馈</title>
<style>body{{font-family:Arial,'Microsoft YaHei',sans-serif;max-width:760px;margin:40px auto;line-height:1.7;color:#243447}}h1{{color:#155e75}}section{{margin:24px 0;padding:18px;background:#f8fafc;border-left:4px solid #0891b2}}dt{{font-weight:bold}}dd{{margin-bottom:8px}}</style></head>
<body><h1>学生代码诊断反馈</h1><section><dl>
<dt>学生姓名</dt><dd>{html.escape(submission.student_name)}</dd>
<dt>题目</dt><dd>{html.escape(task.task.title)}（{html.escape(task.task.problem_id)}）</dd>
<dt>判题结果</dt><dd>{html.escape(submission.status)}</dd>
<dt>通过情况</dt><dd>{submission.passed_count}/{submission.total_count}</dd>
</dl></section><section><h2>错误分析</h2><p>{html.escape(analysis)}</p><ul>{evidence}</ul></section>
<section><h2>改进建议</h2><p>{html.escape(suggestion)}</p></section></body></html>"""


def _prepare_directory(output_dir: str | Path | None) -> Path:
    directory = Path(output_dir) if output_dir else default_data_directory() / "exports"
    try:
        directory.mkdir(parents=True, exist_ok=True)
    except PermissionError as exc:
        raise ExportPermissionError("无法写入导出目录。") from exc
    except OSError as exc:
        raise ReportExportError("报告生成失败。") from exc
    return directory


def _report_stem(task: HistoricalTask) -> str:
    timestamp = task.task.created_at.astimezone().strftime("%Y%m%d_%H%M%S")
    return f"{_safe_filename(task.task.title)}_{timestamp}"


def _safe_filename(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", value).strip(" .")
    return (cleaned or "report")[:80]


def _style_sheet(sheet) -> None:
    sheet.sheet_view.showGridLines = False
    header = sheet[1]
    for cell in header:
        cell.fill = PatternFill("solid", fgColor="0F766E")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for column in range(1, sheet.max_column + 1):
        width = max(
            len(str(sheet.cell(row=row, column=column).value or ""))
            for row in range(1, sheet.max_row + 1)
        )
        sheet.column_dimensions[get_column_letter(column)].width = min(max(width + 3, 12), 36)
