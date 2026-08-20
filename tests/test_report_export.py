from datetime import datetime, timezone

from openpyxl import load_workbook

from models.history import (
    DiagnosisRecord,
    ExplanationRecord,
    HistoricalStudent,
    HistoricalTask,
    SubmissionRecord,
    TaskRecord,
)
from services.report_export import export_class_report, export_student_feedback


def make_task() -> HistoricalTask:
    task = TaskRecord("task-1", "数组训练", "array-1", datetime(2026, 8, 19, tzinfo=timezone.utc), 2)
    first_submission = SubmissionRecord(
        "submission-1", "task-1", "张三", "WA", 8, 10,
        "tasks/task-1/students/a.cpp", "tasks/task-1/results/a.json"
    )
    second_submission = SubmissionRecord(
        "submission-2", "task-1", "李四", "AC", 10, 10,
        "tasks/task-1/students/b.cpp", "tasks/task-1/results/b.json"
    )
    diagnosis = DiagnosisRecord(
        None, "submission-1", "boundary_error", "边界问题", "最大数据失败", 0.8,
        ("小数据通过", "最大数据失败"), (), ()
    )
    explanation = ExplanationRecord(
        None, "submission-1", "SUCCESS", "边界处理不足。", "请检查最大输入。", "基于规则。"
    )
    return HistoricalTask(
        task,
        (
            HistoricalStudent(first_submission, diagnosis, explanation, {}),
            HistoricalStudent(second_submission, None, None, {}),
        ),
    )


def test_xlsx_report_generation_and_content(tmp_path) -> None:
    path = export_class_report(make_task(), output_dir=tmp_path)
    workbook = load_workbook(path, data_only=False)

    assert path.is_file()
    assert workbook.sheetnames == ["总体统计", "学生结果", "错误类型统计"]
    assert workbook["总体统计"]["B5"].value == 2
    assert workbook["总体统计"]["B6"].value == 1
    assert workbook["学生结果"]["A2"].value == "张三"
    assert workbook["学生结果"]["E2"].value == 0.8
    assert workbook["错误类型统计"]["A2"].value == "boundary_error"


def test_student_feedback_html_generation(tmp_path) -> None:
    task = make_task()
    path = export_student_feedback(task, "submission-1", output_dir=tmp_path)
    content = path.read_text(encoding="utf-8")

    assert path.suffix == ".html"
    assert "张三" in content
    assert "数组训练" in content
    assert "WA" in content
    assert "最大数据失败" in content
    assert "请检查最大输入" in content
